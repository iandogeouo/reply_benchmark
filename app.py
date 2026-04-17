import os
import json
import concurrent.futures
from flask import Flask, request, jsonify, send_from_directory
from groq import Groq
from google import genai
from dotenv import load_dotenv
from google.genai import types
import prompts
import time
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

load_dotenv()
gemini_client = genai.Client(api_key=os.getenv('Gemini_API_KEY'))
app    = Flask(__name__, static_folder='.')
groq_client = Groq(api_key=os.getenv('GROQ_API_KEY'))


# ── Groq call (取消註解並確保回傳格式穩定) ───────────────────────────
def call_groq(prompt, model):
    response = groq_client.chat.completions.create(
        model=model,
        messages=[{'role': 'user', 'content': prompt}],
        max_tokens=1000,
        response_format={"type": "json_object"},
        temperature=0,
    )
    text = response.choices[0].message.content
    return json.loads(text)

# ── Gemini call ───────────────────────────────────────────────
def call_gemini(prompt, model):
    response = gemini_client.models.generate_content(
        model=model,
        contents=prompt,
        config=types.GenerateContentConfig(
        temperature=0,
    )
    )
    text = response.candidates[0].content.parts[-1].text
    # 清理 Markdown 代碼塊
    clean = text.replace('```json', '').replace('```', '').strip()
    return json.loads(clean)


# ── 路由：自動分流 ─────────────────────────────────────────────
def get_model_response(prompt, model_name, retries=3):
    last_err = None
    for i in range(retries):
        try:
            if "gemma" in model_name.lower():
                return call_gemini(prompt, model_name)
            else:
                return call_groq(prompt, model_name)
        except json.JSONDecodeError as e:
            last_err = e
            if i < retries - 1:
                time.sleep(1 * (i + 1))
                continue
            raise
        except Exception as e:
            raise  # 非 JSON 錯誤不 retry，直接拋
    raise last_err

# ── Serve static files ─────────────────────────────────────────
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/<path:path>')
def static_files(path):
    return send_from_directory('.', path)


# ── Evaluate endpoint ──────────────────────────────────────────
@app.route('/api/evaluate', methods=['POST'])
def evaluate():
    data       = request.get_json()
    petition   = data.get('petition', '').strip()
    civil      = data.get('civil', '').strip()
    reply      = data.get('reply', '').strip()
    model      = data.get('model', 'llama-3.3-70b-versatile')
    dimensions = data.get('dimensions', ['completeness', 'fidelity', 'tone'])

    if not petition or not reply:
        return jsonify({'error': '缺少必要欄位'}), 400
    if 'fidelity' in dimensions and not civil:
        return jsonify({'error': '評估忠實性需要提供公務員想回答的內容'}), 400
    if not dimensions:
        return jsonify({'error': '請至少選擇一個評估維度'}), 400

    # 依勾選的維度建立要執行的任務
    tasks = {}
    if 'completeness' in dimensions:
        tasks['completeness'] = prompts.completeness(petition, reply)
    if 'fidelity' in dimensions:
        tasks['fidelity'] = prompts.fidelity(petition, civil, reply)
    if 'tone' in dimensions:
        tasks['tone'] = prompts.tone(reply)

    # 平行呼叫
    results = {}
    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = {
            executor.submit(get_model_response, prompt, model): dim
            for dim, prompt in tasks.items()
        }
        for future in concurrent.futures.as_completed(futures):
            dim = futures[future]
            try:
                results[dim] = future.result()
            except json.JSONDecodeError as e:
                return jsonify({'error': f'{dim}: 模型回傳格式錯誤', 'detail': str(e)}), 422
            except Exception as e:
                import traceback
                traceback.print_exc() 
                return jsonify({'error': f'{dim}: {str(e)}'}), 500

    return jsonify(results)


# ── 儲存評測結果 ─────────────────────────────────────────


RECORDS_FOLDER = 'records'
os.makedirs(RECORDS_FOLDER, exist_ok=True)

@app.route('/api/save', methods=['POST'])
def save_record():
    data = request.get_json()
    mode = data.get('mode', 'single')
    rows = data.get('rows', [])
    model = data.get('model', '')

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{timestamp}_{mode}.xlsx"
    filepath = os.path.join(RECORDS_FOLDER, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = '評測結果'
    ws.sheet_view.showGridLines = False  # 隱藏格線，更乾淨

    # ── 顏色定義 ──
    COLOR_HEADER_BG  = '1E293B'   # 深藍灰
    COLOR_HEADER_FG  = 'FFFFFF'
    COLOR_SCORE_BG   = 'EFF6FF'   # 淡藍（分數欄）
    COLOR_REASON_BG  = 'F8FAFC'   # 極淡灰（原因欄）
    COLOR_ROW_ALT    = 'F1F5F9'   # 交替行
    COLOR_ROW_WHITE  = 'FFFFFF'
    COLOR_SCORE_5    = '16A34A'   # 綠
    COLOR_SCORE_4    = '65A30D'   # 黃綠
    COLOR_SCORE_3    = 'D97706'   # 橘
    COLOR_SCORE_2    = 'DC2626'   # 紅
    COLOR_BORDER     = 'CBD5E1'

    thin = Side(style='thin', color=COLOR_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def score_color(s):
        if s is None: return None
        if s >= 5: return COLOR_SCORE_5
        if s >= 4: return COLOR_SCORE_4
        if s >= 3: return COLOR_SCORE_3
        return COLOR_SCORE_2

    # ── 標頭 ──
    headers = [
        ('編號', 6), ('時間', 18), ('評估模型', 15),
        ('陳情內容', 45), ('公務員輸入', 35), ('擬答內容', 45),
        ('完整性\n分數', 9), ('完整性原因', 55),
        ('忠實性\n分數', 9), ('忠實性原因', 55),
        ('語調\n分數', 9),  ('語調原因', 55),
    ]

    for col, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = PatternFill('solid', start_color=COLOR_HEADER_BG)
        cell.font      = Font(bold=True, color=COLOR_HEADER_FG, name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 36

    # ── 資料 ──
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    score_cols   = {7, 9, 11}   # 分數欄 index（1-based）
    reason_cols  = {8, 10, 12}  # 原因欄

    for i, row in enumerate(rows, 2):
        r       = row.get('results', {})
        is_alt  = (i % 2 == 0)
        row_bg  = COLOR_ROW_ALT if is_alt else COLOR_ROW_WHITE

        scores = {
            7:  r.get('completeness', {}).get('score'),
            9:  r.get('fidelity',     {}).get('score'),
            11: r.get('tone',         {}).get('score'),
        }
        reasons = {
            8:  r.get('completeness', {}).get('reason', ''),
            10: r.get('fidelity',     {}).get('reason', ''),
            12: r.get('tone',         {}).get('reason', ''),
        }

        values = [
            i - 1, now_str, model,
            row.get('petition', ''), row.get('civil', ''), row.get('reply', ''),
            scores[7], reasons[8],
            scores[9], reasons[10],
            scores[11], reasons[12],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = border
            cell.font   = Font(name='Arial', size=10)

            if col in score_cols:
                s = scores[col]
                sc = score_color(s)
                cell.fill      = PatternFill('solid', start_color=COLOR_SCORE_BG)
                cell.font      = Font(name='Arial', size=11, bold=True,
                                      color=sc if sc else '334155')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col in reason_cols:
                cell.fill      = PatternFill('solid', start_color=COLOR_REASON_BG)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.font      = Font(name='Arial', size=9, color='475569')
            else:
                cell.fill      = PatternFill('solid', start_color=row_bg)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.row_dimensions[i].height = 80

    # ── 凍結第一列 ──
    ws.freeze_panes = 'A2'

    wb.save(filepath)

    # 同時存 JSON 供快速讀取
    json_rows = []
    for i, row in enumerate(rows):
        r = row.get('results', {})
        json_rows.append({
            '編號':       i + 1,
            '時間':       now_str,
            '評估模型':   model,
            '陳情內容':   row.get('petition', ''),
            '公務員輸入': row.get('civil', ''),
            '擬答內容':   row.get('reply', ''),
            '完整性\n分數': r.get('completeness', {}).get('score'),
            '完整性原因':  r.get('completeness', {}).get('reason', ''),
            '忠實性\n分數': r.get('fidelity', {}).get('score'),
            '忠實性原因':  r.get('fidelity', {}).get('reason', ''),
            '語調\n分數':  r.get('tone', {}).get('score'),
            '語調原因':    r.get('tone', {}).get('reason', ''),
        })
    json_path = os.path.join(RECORDS_FOLDER, filename.replace('.xlsx', '.json'))
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(json_rows, f, ensure_ascii=False)

    def avg_score(key):
        vals = [r.get('results', {}).get(key, {}).get('score') for r in rows]
        vals = [v for v in vals if v is not None]
        return round(sum(vals) / len(vals), 1) if vals else None

    meta_path = os.path.join(RECORDS_FOLDER, 'metadata.json')
    metadata  = []
    if os.path.exists(meta_path):
        with open(meta_path, 'r', encoding='utf-8') as f:
            metadata = json.load(f)

    metadata.append({
        'filename':        filename,
        'mode':            mode,
        'model':           model,
        'count':           len(rows),
        'timestamp':       datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'avg_completeness': avg_score('completeness'),
        'avg_fidelity':    avg_score('fidelity'),
        'avg_tone':        avg_score('tone'),
    })

    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)
    return jsonify({'status': 'ok', 'filename': filename})

# ── 讀取紀錄 ───────────────────────────────────────
@app.route('/api/records', methods=['GET'])
def list_records():
    meta_path = os.path.join(RECORDS_FOLDER, 'metadata.json')
    if not os.path.exists(meta_path):
        return jsonify([])
    with open(meta_path, 'r', encoding='utf-8') as f:
        return jsonify(json.load(f))

@app.route('/api/records/<filename>', methods=['GET'])
def get_record(filename):
    # 優先讀 JSON（快）
    json_path = os.path.join(RECORDS_FOLDER, filename.replace('.xlsx', '.json'))
    if os.path.exists(json_path):
        with open(json_path, 'r', encoding='utf-8') as f:
            return jsonify(json.load(f))

    # 舊紀錄 fallback：解析 xlsx（慢）
    filepath = os.path.join(RECORDS_FOLDER, filename)
    if not os.path.exists(filepath):
        return jsonify({'error': '找不到檔案'}), 404

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return jsonify(rows)

@app.route('/api/records/<filename>', methods=['DELETE'])
def delete_record(filename):
    meta_path = os.path.join(RECORDS_FOLDER, 'metadata.json')
    if os.path.exists(meta_path):
        with open(meta_path, 'r', encoding='utf-8') as f:
            metadata = json.load(f)
        metadata = [m for m in metadata if m['filename'] != filename]
        with open(meta_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)

    for ext in ['.xlsx', '.json']:
        path = os.path.join(RECORDS_FOLDER, filename.replace('.xlsx', ext))
        if os.path.exists(path):
            os.remove(path)

    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
